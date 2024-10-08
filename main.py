import os
import os.path
import sys

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Border, Side, Alignment
from PyQt6.QtWidgets import QApplication
from PyQt6 import QtWidgets

from styles import Styles
from current_books import AgroBookSunflower, AgroBookRapeSeed, CornBook
from settings import DATA_DIR, FILE_NAME
from ui.main_window import AgroMainWindow, DataProcessor, CropsRadio


def get_file_name_by_path(file_path):
    _, file_name = os.path.split(file_path)
    return file_name


def get_wb(input_file):
    try:
        workbook = load_workbook(input_file)
    except Exception:
        file_name = get_file_name_by_path(input_file)
        raise ValueError(f'Невдалось відкрити файл "{file_name}"')
    return workbook


def get_ws(wb, input_file):
    sheet_name = 'Whole data'
    if sheet_name not in wb.sheetnames:
        file_name = get_file_name_by_path(input_file)
        raise ValueError(f'Відсутній лист "{sheet_name}" у файлі "{file_name}"')
    ws = wb[sheet_name]
    return ws


def open_book(input_file):
    wb = get_wb(input_file)
    ws = get_ws(wb, input_file)
    return ws


def get_columns(input_ws):
    columns = {}
    for row in input_ws:
        for col in row:
            columns[col.column_letter] = col.value
        break
    return columns


def process_file(agro_book, input_ws):
    columns = get_columns(input_ws)
    agro_book.process_input_ws(input_ws=input_ws, columns=columns)
    for i, row in enumerate(input_ws.rows):
        if i == 0:
            continue
        agro_book.write_from_row(row, columns)


def fill_book(agro_book, file_list):
    data_processor = DataProcessor()
    for index, input_file in enumerate(file_list):
        print(f'==Обробка файлу {get_file_name_by_path(input_file)}==')
        try:
            input_ws = open_book(input_file)
            process_file(agro_book, input_ws)
        except Exception as e:
            print(e, file=sys.stderr)

        progress = int(((index + 1) * 100) / len(file_list))
        data_processor.update_progress(progress)


def file_exists(file_path):
    if not os.path.exists(file_path):
        raise ValueError(f'Схоже, що файла {file_path} не існує')


def its_file(file_path):
    if not os.path.isfile(file_path):
        raise ValueError(f'{file_path} має бути файлом')


def init_file_name_and_data_dir(file_path, dir_path):
    global FILE_NAME
    global DATA_DIR
    if file_path is not None:
        file_exists(file_path)
        its_file(file_path)

        DATA_DIR, FILE_NAME = os.path.split(file_path)
    else:
        DATA_DIR = dir_path


def main(*, file_list, crops, file_path=None, dir_path=None):
    if file_path is None and dir_path is None:
        raise ValueError("Має бути вказано або ім'я файла або директорія")

    if crops is None:
        raise ValueError('Ви не вибрали культуру')

    books_mapping = {
        CropsRadio.sunflower.name: AgroBookSunflower,
        CropsRadio.rapeseed.name: AgroBookRapeSeed,
        CropsRadio.corn.name: CornBook
    }

    book = books_mapping.get(crops)

    init_file_name_and_data_dir(file_path, dir_path)
    agro_book = book(file_path)
    fill_book(agro_book, file_list)
    agro_book.save_book(FILE_NAME, DATA_DIR)


if __name__ == '__main__':
    import resources_rc

    if getattr(sys, 'frozen', False):
        application_path = sys._MEIPASS
    else:
        application_path = os.path.dirname(os.path.abspath(__file__))

    style_path = os.path.join(application_path, 'ui/style.gss')

    with open(style_path) as style_file:
        style_data = style_file.read()

    app = QApplication(sys.argv)
    app.setStyleSheet(style_data)

    window = AgroMainWindow()
    window.show()

    sys.exit(app.exec())
