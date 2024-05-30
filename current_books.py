import sys
from collections import defaultdict
from openpyxl import Workbook, load_workbook
from styles import Styles


class BaseBook:
    sheet_name = 'aggregate_yield'
    column_lib = {
        'Область': {'type': 'handle', 'width': 19.35},
        'Район': {'type': 'simple', 'width': 17.36},
        'Господарство': {'type': 'handle', 'width': 22.80},
        'GPS-координати поля': {'type': 'handle', 'width': 28.53},
        'COMPANY': {'type': 'simple', 'width': 15.41},
        'HYBRIDS': {'type': 'simple', 'width': 20.24},
        'Обробіток грунту': {'type': 'simple', 'width': 15.77},
        'Тип грунту': {'type': 'simple', 'width': 15.77},
        'Густота на момент\nзбирання, тис/га': {'type': 'simple', 'width': 19.92},
        'Кіл-ть\nрядків': {'type': 'simple', 'width': 8.57},
        'Довжина\nділянки': {'type': 'simple', 'width': 8.57},
        'Площа, га': {'type': 'handle', 'width': 10.0, 'style': {'number_format': '0.0000'}},
        'Вага з\nділянки, кг': {'type': 'simple', 'width': 10.65},
        'YIELD,\nBUNKER WEIGHT (q/ha)\nБункерна вага': {'type': 'handle', 'width': 20.53, 'style': {'number_format': '0.00'}},
        'Harvesting moisture,\n% Вологість': {'type': 'simple', 'width': 18.29},
        'Вологість на час\nзбирання, %': {'type': 'simple', 'width': 18.29},
        'Re-calculation\nof yield at basis\nmoisture (7 %) (UA)': {'type': 'handle', 'width': 19.80, 'style': {'number_format': '0.00'}},
        'Re-calculation\nof yield at basis\nmoisture (8 %) (F)': {'type': 'handle', 'width': 19.80, 'style': {'number_format': '0.00'}},
        'Урожайність у\nперерахунку на базисну\nвологість (9 %) (UA)': {'type': 'handle', 'width': 22.47, 'style': {'number_format': '0.00'}},
        'Попередник': {'type': 'simple', 'width': 16.71},
        'Дата посіву': {'type': 'simple', 'width': 11.10},
        'Дата збирання': {'type': 'simple', 'width': 13.74},
        'ПІБ менеджера,\nщо створив протокол': {'type': 'simple', 'width': 22.03},
        'Тип досліду\n(Demo/SBS/Strip)': {'type': 'simple', 'width': 15.33},
        'Ширина\nміжряддя': {'type': 'simple', 'width': 16.63},
        'Ширина\nділянки': {'type': 'simple', 'width': 16.63},
        'Коментарі': {'type': 'handle', 'width': 35.20},
        'БР': {'type': 'handle', 'width': 8.27},
        'Рік': {'type': 'simple', 'width': 9.37},
        'ФАО': {'type': 'handle', 'width': 20.47},
        'Урожайність  (в перерахунку на вологість зерна 14%), ц/га': {'type': 'simple', 'width': 15.20},
        'Коеф. урож SY': {'type': 'handle', 'width': 15.20},
        'Коеф. урож SY+конк.': {'type': 'handle', 'width': 15.20},
    }

    def __init__(self, file_path=None):
        self.wb = None
        self.ws = None
        self.num_rows = 0
        self.styles = Styles()
        self.columns = self.init_columns()

        self.init_workbook(file_path)

    def get_column_from_lib(self, title, letter, func=None, rel_col=None):
        column_setting = self.column_lib.get(title)
        if column_setting is None:
            raise ValueError(f'В бібліотеці колонок не знайдено {title}')

        column_type = column_setting['type']
        if column_type == 'handle':
            column_setting['func'] = func
        elif column_type == 'simple':
            column_setting['rel_col'] = rel_col

        column_setting['letter'] = letter

        return column_setting

    def init_columns(self):
        return {'NN': {'type': 'handle', 'func': self.line_number, 'letter': 'A', 'width': 6.54}}

    def init_workbook(self, file_path):
        if file_path is None:
            self.wb = Workbook()
            self.ws = self.get_ws()
            self.num_rows = 1

            self.init_default_columns()
        else:
            try:
                self.wb = load_workbook(file_path)
            except Exception:
                raise RuntimeError(f'Не вдалось відкрити файл {file_path}')

            self.ws = self.get_ws_from_file(self.wb)
            self.num_rows = self.ws.max_row

            self.check_file_structure()

    def get_ws(self):
        ws = self.wb.active
        ws.title = self.sheet_name
        return ws

    def get_ws_from_file(self, wb):
        if self.sheet_name not in wb.sheetnames:
            raise ValueError(f'Неправильна структура файла. Відсутній лист {self.sheet_name}')

        ws = wb.get_sheet_by_name(self.sheet_name)
        return ws

    def check_file_structure(self):
        default_column_list = list(self.columns.keys())
        first_row = next(self.ws.iter_rows())
        file_column = [cell.value for cell in first_row]

        # the number of elements must be the same
        not_enough_items = max(len(default_column_list) - len(file_column), 0)
        file_column.extend([''] * not_enough_items)

        for i, col_name in enumerate(default_column_list):
            if col_name != file_column[i]:
                raise ValueError(f'Неправильна структура файла. Проблема в назві колонки {i+1}')

    def init_default_columns(self):
        for column_name, data in self.columns.items():
            cell = self.ws[f'{data["letter"]}1']
            cell.value = column_name
            self.apply_title_styles(cell)
            self.adjust_column_width(data["letter"], data['width'])
        self.ws.row_dimensions[1].height = 46.49

    def adjust_column_width(self, column_letter, width):
        self.ws.column_dimensions[column_letter].width = width

    def apply_title_styles(self, cell):
        self.apply_styles(cell)
        cell.font = self.styles.font_bold

    def apply_styles(self, cell):
        cell.font = self.styles.font
        cell.border = self.styles.border
        cell.alignment = self.styles.wrap_text()

    def row_to_dict(self, row, column):
        row_dict = {}
        for cell in row:
            row_dict[column[cell.column_letter]] = cell.value
        return row_dict

    def write_data(self, row_dict):
        self.num_rows += 1

        for title, data in self.columns.items():
            cell_coordinate = f'{data["letter"]}{self.num_rows}'
            cell = self.ws[cell_coordinate]

            if data['type'] == 'handle' and data['func']:
                method = data['func']

                try:
                    val = method(row_dict)
                except Exception as e:
                    print(f'Помилка при розрахунку {cell_coordinate}. {e}', file=sys.stderr)
                    val = None

                cell.value = val

            elif data['type'] == 'simple':
                cell.value = row_dict.get(data['rel_col'], None)

            self.apply_styles(cell)

            if data.get('style'):
                style = data.get('style')
                for key, value in style.items():
                    setattr(cell, key, value)

    def write_from_row(self, file_row, file_columns):
        row_dict = self.row_to_dict(file_row, file_columns)
        self.write_data(row_dict)

    def save_book(self, file_name, file_dir):
        self.wb.save(f'{file_dir}/{file_name}')

    def line_number(self, row_dict):
        return self.num_rows - 1

    def calculate_the_area_by_formula(self):
        pass

    def format_coordinates(self, coordinate):
        if coordinate is None:
            return ''

        coord = str(coordinate)
        coord.replace(',', '.')
        return coord

    def gps_coordinates(self, row_dict):
        longitude = row_dict.get('Longitude', None)
        longitude = self.format_coordinates(longitude)

        latitude = row_dict.get('Latitude', None)
        latitude = self.format_coordinates(latitude)

        if longitude == '' or latitude == '':
            return None

        return f'{latitude},{longitude}'

    def process_input_ws(self, *args, **kwargs):
        pass


class OilSeedCropBook(BaseBook):
    def state(self, row_dict):
        rel_col = 'State'
        value = row_dict.get(rel_col, '')
        if value is None:
            return None
        return value.replace('область', '').strip()

    def household(self, row_dict):
        value_row = row_dict.get('Custom trial name', '')
        if value_row is None:
            return None
        value = value_row.split(',')
        if len(value) >= 4:
            return value[3]
        return value_row

    def get_not_none_value(self, col_name, row_dict):
        value = row_dict.get(col_name, None)
        if value is None:
            raise ValueError(f'Не вдалось отримати дані колонки "{col_name}"')
        return value

    def get_float_value(self, value):
        try:
            value = float(value)
        except ValueError:
            raise ValueError(
                f'Дані {value} не можуть бути конвертовані в число')
        return value

    def calculate_the_area_from_hrvar(self, row_dict):
        hrvar = self.get_not_none_value('Hrvar', row_dict)
        hrvar = self.get_float_value(hrvar)
        return round((hrvar / 10000), 4)

    def area_hectares(self, row_dict):
        number_of_rows = row_dict.get('Number of rows', None)
        length_of_plot = row_dict.get('Plot Length', None)

        if number_of_rows is None or length_of_plot is None:
            area = self.calculate_the_area_from_hrvar(row_dict)
        else:
            area = self.calculate_the_area_by_formula()

        return area

    def yield_bunker_weight_field(self, row_dict):
        return f'=M{self.num_rows}/(L{self.num_rows}*100)'

    def yield_recalculation_field_7(self, row_dict):
        return f'=N{self.num_rows}*((100-O{self.num_rows})/93)'

    def yield_recalculation_field_8(self, row_dict):
        return f'=N{self.num_rows}*((100-O{self.num_rows})/92)'

    def calculate_yield_bunker_weight(self, row_dict):
        return f'=Q{self.num_rows}/(P{self.num_rows}*100)'

    def yield_recalculation_field_9(self, row_dict):
        return f'=R{self.num_rows}*((100-S{self.num_rows})/91)'


class AgroBookSunflower(OilSeedCropBook):
    def init_columns(self):
        columns = super().init_columns()

        columns.update({
            'Область': self.get_column_from_lib('Область', func=self.state, letter='B'),
            'Район': self.get_column_from_lib('Район', rel_col='City', letter='C'),
            'Господарство': self.get_column_from_lib('Господарство', func=self.household, letter='D'),
            'GPS-координати поля': self.get_column_from_lib('GPS-координати поля', func=self.gps_coordinates, letter='E'),
            'COMPANY': self.get_column_from_lib('COMPANY', rel_col='Hybrid Company Name', letter='F'),
            'HYBRIDS': self.get_column_from_lib('HYBRIDS', rel_col='Hybrid Name', letter='G'),
            'Обробіток грунту': self.get_column_from_lib('Обробіток грунту', rel_col='PTL_C', letter='H'),
            'Густота на момент\nзбирання, тис/га': self.get_column_from_lib('Густота на момент\nзбирання, тис/га', rel_col='HAVPN', letter='I'),
            'Кіл-ть\nрядків': self.get_column_from_lib('Кіл-ть\nрядків', rel_col='Number of rows', letter='J'),
            'Довжина\nділянки': self.get_column_from_lib('Довжина\nділянки', rel_col='Plot Length', letter='K'),
            'Площа, га': self.get_column_from_lib('Площа, га', func=self.area_hectares, letter='L'),
            'Вага з\nділянки, кг': self.get_column_from_lib('Вага з\nділянки, кг', rel_col='GWTPN', letter='M'),
            'YIELD,\nBUNKER WEIGHT (q/ha)\nБункерна вага': self.get_column_from_lib('YIELD,\nBUNKER WEIGHT (q/ha)\nБункерна вага', func=self.yield_bunker_weight_field, letter='N'),
            'Harvesting moisture,\n% Вологість': self.get_column_from_lib('Harvesting moisture,\n% Вологість', rel_col='GMSTP', letter='O'),
            'Re-calculation\nof yield at basis\nmoisture (7 %) (UA)': self.get_column_from_lib('Re-calculation\nof yield at basis\nmoisture (7 %) (UA)', func=self.yield_recalculation_field_7, letter='P'),
            'Re-calculation\nof yield at basis\nmoisture (8 %) (F)': self.get_column_from_lib('Re-calculation\nof yield at basis\nmoisture (8 %) (F)', func=self.yield_recalculation_field_8, letter='Q'),
            'Попередник': self.get_column_from_lib('Попередник', rel_col='Previous Crop', letter='R'),
            'Дата посіву': self.get_column_from_lib('Дата посіву', rel_col='Date of Planting', letter='S'),
            'Дата збирання': self.get_column_from_lib('Дата збирання', rel_col='Date of Harvest', letter='T'),
            'ПІБ менеджера,\nщо створив протокол': self.get_column_from_lib('ПІБ менеджера,\nщо створив протокол', rel_col='Username', letter='U'),
            'Тип досліду\n(Demo/SBS/Strip)': self.get_column_from_lib('Тип досліду\n(Demo/SBS/Strip)', rel_col='Trial type', letter='V'),
            'Ширина\nміжряддя': self.get_column_from_lib('Ширина\nміжряддя', rel_col='Row spacing', letter='W'),
            'Коментарі': self.get_column_from_lib('Коментарі', func=None, letter='X')
        })

        return columns

    def calculate_the_area_by_formula(self):
        return f'=J{self.num_rows}*0.7*K{self.num_rows}/10000'


class AgroBookRapeSeed(OilSeedCropBook):
    def init_columns(self):
        columns = super().init_columns()

        sales_manager_full_name = self.get_column_from_lib('ПІБ менеджера,\nщо створив протокол', rel_col='Username', letter='U')
        sales_manager_full_name['width'] = 26.48

        soil_processing = self.get_column_from_lib('Обробіток грунту', rel_col='PTL_C', letter='I')
        soil_processing['width'] = 20.63

        columns.update({
            'Область': self.get_column_from_lib('Область', func=self.state, letter='B'),
            'Район': self.get_column_from_lib('Район', rel_col='City', letter='C'),
            'Господарство': self.get_column_from_lib('Господарство', func=self.household, letter='D'),
            'GPS-координати поля': self.get_column_from_lib('GPS-координати поля', func=self.gps_coordinates, letter='E'),
            'Компанія': self.get_column_from_lib('COMPANY', rel_col='Hybrid Company Name', letter='F'),
            'Гібрид': self.get_column_from_lib('HYBRIDS', rel_col='Hybrid Name', letter='G'),
            'Попередник': self.get_column_from_lib('Попередник', rel_col='Previous Crop', letter='H'),
            'Обробіток грунту': soil_processing,
            'Тип грунту': self.get_column_from_lib('Тип грунту', rel_col='USDAS', letter='J'),
            'Дата посіву': self.get_column_from_lib('Дата посіву', rel_col='Date of planting', letter='K'),
            'Дата збирання': self.get_column_from_lib('Дата збирання', rel_col='Date of Harvest', letter='L'),
            'Ширина\nміжряддя,\nсм': self.get_column_from_lib('Ширина\nміжряддя', func=None, letter='M'),
            'Довжина\nділянки,\nм': self.get_column_from_lib('Довжина\nділянки', rel_col='Plot Length', letter='N'),
            'Ширина\nділянки,\nм': self.get_column_from_lib('Ширина\nділянки', rel_col='Plot Width', letter='O'),
            'Площа, га': self.get_column_from_lib('Площа, га', func=self.area_hectares, letter='P'),
            'Вага з\nділянки, кг': self.get_column_from_lib('Вага з\nділянки, кг', rel_col='USY', letter='Q'),
            'YIELD,\nBUNKER WEIGHT (q/ha)\nБункерна вага': self.get_column_from_lib('YIELD,\nBUNKER WEIGHT (q/ha)\nБункерна вага', func=self.calculate_yield_bunker_weight, letter='R'),
            'Вологість на час\nзбирання, %': self.get_column_from_lib('Вологість на час\nзбирання, %', rel_col='HSH', letter='S'),
            'Урожайність у\nперерахунку на базисну\nвологість (9 %) (UA)': self.get_column_from_lib('Урожайність у\nперерахунку на базисну\nвологість (9 %) (UA)', func=self.yield_recalculation_field_9, letter='T'),
            'ПІБ Менеджер з продажів': sales_manager_full_name,
            'Тип досліду\n(Demo/SBS/Strip)': self.get_column_from_lib('Тип досліду\n(Demo/SBS/Strip)', rel_col='Trial type', letter='V'),
        })

        return columns

    def calculate_the_area_by_formula(self):
        return f'=(N{self.num_rows}*O{self.num_rows})/10000'

    def area_hectares(self, row_dict):
        width_of_plot = row_dict.get('Plot Width', None)
        length_of_plot = row_dict.get('Plot Length', None)

        if width_of_plot is None or length_of_plot is None:
            area = self.calculate_the_area_from_hrvar(row_dict)
        else:
            area = self.calculate_the_area_by_formula()

        return area

    def household(self, row_dict):
        value_row = row_dict.get('Custom trial name', None)
        return value_row


class CornBook(BaseBook):
    yield_field = 'YGSMN'
    company_field = 'Hybrid Company Name'

    def __init__(self, file_path=None):
        super().__init__(file_path=file_path)

        self.average_yield_sy = None
        self.average_yield_with_competitors = None

    def init_columns(self):
        columns = super().init_columns()
        columns.update({
            'БР': self.get_column_from_lib('БР', func=None, letter='B'),
            'Виробник': self.get_column_from_lib('COMPANY', rel_col=self.company_field, letter='C'),
            'Попередник\n(культура)': self.get_column_from_lib('Попередник', rel_col='Previous Crop', letter='D'),
            'Рік': self.get_column_from_lib('Рік', rel_col='Year', letter='E'),
            'Область': self.get_column_from_lib('Область', func=self.state, letter='F'),
            'Район': self.get_column_from_lib('Район', func=None, letter='G'),
            'Локація': self.get_column_from_lib('Господарство', func=self.household, letter='H'),
            'GPS-координати поля': self.get_column_from_lib('GPS-координати поля', func=self.gps_coordinates, letter='I'),
            'Гібрид': self.get_column_from_lib('HYBRIDS', rel_col='Hybrid Name', letter='J'),
            'ФАО': self.get_column_from_lib('ФАО', func=None, letter='K'),
            'Вологість зерна під час збирання %': self.get_column_from_lib('Harvesting moisture,\n% Вологість', rel_col='GMSTP', letter='L'),
            'Урожайність  (в перерахунку на вологість зерна 14%), ц/га': self.get_column_from_lib('Урожайність  (в перерахунку на вологість зерна 14%), ц/га', rel_col=self.yield_field, letter='M'),
            'Коеф. урож SY': self.get_column_from_lib('Коеф. урож SY', func=self.crop_yield_coefficient_sy, letter='N'),
            'Коеф. урож SY+конк.': self.get_column_from_lib('Коеф. урож SY+конк.', func=self.crop_yield_coefficient_with_competitors, letter='O'),
            'Дата посіву': self.get_column_from_lib('Дата посіву', rel_col='Date of Planting', letter='P'),
            'Дата збирання': self.get_column_from_lib('Дата збирання', rel_col='Date of Harvest', letter='Q'),
            'ПІБ менеджера,\nщо створив протокол': self.get_column_from_lib('ПІБ менеджера,\nщо створив протокол', rel_col='Username', letter='R'),
            'Тип досліду\n(Demo/SBS/Strip)': self.get_column_from_lib('Тип досліду\n(Demo/SBS/Strip)', rel_col='Trial type', letter='S'),
        })

        return columns

    def state(self, row_dict):
        value_row = row_dict.get('State', None)
        return value_row

    def household(self, row_dict):
        value_row = row_dict.get('Custom trial name', None)
        return value_row

    def crop_yield_coefficient_sy(self, row_dict):
        if self.average_yield_sy is None:
            return None

        return f'=M{self.num_rows}/{self.average_yield_sy}'

    def crop_yield_coefficient_with_competitors(self, row_dict):
        if self.average_yield_with_competitors is None:
            return None

        return f'=M{self.num_rows}/{self.average_yield_with_competitors}'

    def process_input_ws(self, *args, **kwargs):
        self.average_yield_sy = None
        self.average_yield_with_competitors = None

        input_ws = kwargs.get('input_ws')
        columns = kwargs.get('columns')

        if input_ws is None or columns is None:
            raise ValueError('Неможливо розрахувати середню врожайність. Відсутні input_ws і columns')

        yield_letter = None
        company_letter = None

        for k, v in columns.items():
            if v == self.yield_field:
                yield_letter = k

            if v == self.company_field:
                company_letter = k

        if yield_letter is None or company_letter is None:
            raise ValueError('Неможливо розрахувати середню врожайність. Відсутні yield_letter, company_letter')

        self.average_yields(input_ws, yield_letter, company_letter)

    def average_yields(self, input_ws, yield_letter, company_letter):
        average_dict = defaultdict(list)

        for row_number in range(2, input_ws.max_row + 1):

            yield_value = self.get_yield_value(input_ws, row_number, yield_letter)
            company_value = self.get_company_value(input_ws, row_number, company_letter)
            average_dict[company_value].append(yield_value)

        self.average_syngenta(average_dict)
        self.average_with_competitors(average_dict)

    def average_with_competitors(self, average_dict):
        keys_list = list(average_dict.keys())
        list_without_syngenta = list(filter(lambda v: v != 'syngenta', keys_list))
        if not list_without_syngenta:
            return

        common_list = []
        for company, yields_list in average_dict.items():
            common_list.extend(yields_list)

        average = round(sum(common_list) / len(common_list), 2)
        self.average_yield_with_competitors = average

    def average_syngenta(self, average_dict):
        syngenta_list = average_dict.get('syngenta', None)
        if not syngenta_list:
            raise ValueError('Неможливо розрахувати середню врожайність. Відсутні врожаї для Syngenta')

        average = round(sum(syngenta_list) / len(syngenta_list), 2)
        self.average_yield_sy = average

    def get_yield_value(self, input_ws, row_number, yield_letter):
        try:
            value = input_ws[f'{yield_letter}{row_number}'].value
            value = float(value)
            return value
        except Exception:
            return 0.0

    def get_company_value(self, input_ws, row_number, company_letter):
        value = input_ws[f'{company_letter}{row_number}'].value or ''
        return value.lower()

